#!/usr/bin/env python3

"""
Room Item Swap Import Workflow

This workflow processes room item swap spreadsheets to swap out old items with new items in rooms.
The spreadsheet should contain:
- room_sku: The SKU of the room to update
- concatenated_itemqty: Pipe-separated list of old item SKUs with quantities (format: "SKU:QTY|SKU:QTY")
- item_sku: The new item SKU to swap in
- division: Division (FL, SE, TX) or site

The workflow:
1. Parses the spreadsheet
2. Groups swaps by room and division
3. Validates the data
4. Exports each swap to its own JSON file for traceability
5. Executes each swap individually via API (one work request per room)

Note: Each room swap is isolated and gets its own JSON file and work request
for better traceability and failure isolation.
"""

import pandas as pd
import click
import os
import shutil
import json
import re
from rich.console import Console
from rich.progress import Progress
from rich.table import Table
from rich.prompt import Confirm
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dotenv import load_dotenv
from datetime import datetime

from ecatalog_client import (
    ECatalogAPIClient,
    SwapRoomItemsRequest,
    RoomItem,
    OAuthConfig,
)

console = Console()

# Load environment variables
load_dotenv()


class RoomItemSwapImporter:
    """Import and process room item swap data from spreadsheets"""

    def __init__(self, api_client: ECatalogAPIClient):
        self.client = api_client
        self.console = console

    def parse_concatenated_items(self, concatenated_str: str) -> List[Dict[str, any]]:
        """
        Parse concatenated item quantity string into list of items

        Format: "SKU1:QTY1|SKU2:QTY2|..."
        Example: "23056066:1|22256069:1"

        Returns:
            List of dicts with 'sku' and 'quantity' keys
        """
        if not concatenated_str or pd.isna(concatenated_str):
            return []

        items = []
        # Split by pipe character
        parts = str(concatenated_str).split('|')

        for part in parts:
            if ':' in part:
                sku, qty = part.split(':', 1)
                items.append({
                    'sku': sku.strip(),
                    'quantity': int(qty.strip())
                })

        return items

    def normalize_division(self, value: str) -> Optional[str]:
        """
        Normalize division value from various formats

        Args:
            value: Input value (could be "FL", "RTG", "SE", etc.)

        Returns:
            Normalized division code (FL, SE, or TX) or None
        """
        if not value or pd.isna(value):
            return None

        value = str(value).upper().strip()

        # Direct division codes
        if value in ['FL', 'SE', 'TX']:
            return value

        # Site to division mapping
        site_to_division = {
            'RTG': 'FL',  # Rooms To Go -> Florida
            'OTG': 'SE',  # Out There Go -> Southeast
            'KTG': 'TX',  # Kinda Texas Go -> Texas (or adjust as needed)
        }

        return site_to_division.get(value)

    def validate_swap_data(self, swap_data: Dict) -> Tuple[bool, List[str]]:
        """
        Validate a room item swap before execution

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []

        # Required fields
        if not swap_data.get('room_sku'):
            errors.append("Missing room_sku")

        if not swap_data.get('swap_out_items'):
            errors.append("No items to swap out")

        if not swap_data.get('swap_in_items'):
            errors.append("No items to swap in")

        if not swap_data.get('divisions'):
            errors.append("No divisions specified")

        # Validate division codes
        for div in swap_data.get('divisions', []):
            if div not in ['FL', 'SE', 'TX']:
                errors.append(f"Invalid division code: {div}")

        # Validate SKU formats (basic check)
        for item in swap_data.get('swap_out_items', []):
            if not item.get('sku'):
                errors.append("Swap out item missing SKU")
            if not item.get('quantity') or item['quantity'] <= 0:
                errors.append(f"Invalid quantity for swap out item {item.get('sku')}")

        for item in swap_data.get('swap_in_items', []):
            if not item.get('sku'):
                errors.append("Swap in item missing SKU")
            if not item.get('quantity') or item['quantity'] <= 0:
                errors.append(f"Invalid quantity for swap in item {item.get('sku')}")

        return (len(errors) == 0, errors)

    def process_spreadsheet(self, file_path: Path, sheet_name: Optional[str] = None) -> List[Dict]:
        """
        Process room item swap spreadsheet

        Returns:
            List of swap operations to perform
        """
        console.print(f"[cyan]Reading file: {file_path}[/cyan]")

        # Determine file type and read accordingly
        if file_path.suffix.lower() == '.csv':
            df = pd.read_csv(file_path)
        else:
            # For old Excel format, we need to use a different approach
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb.active if sheet_name is None else wb[sheet_name]

                # Convert to list of dicts
                data = []
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data.append(dict(zip(headers, row)))
                df = pd.DataFrame(data)
            except Exception as e:
                console.print(f"[yellow]Warning: Could not read as XLSX, trying XLS format[/yellow]")
                # Fall back to pandas with xlrd
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')

        console.print(f"[green]Loaded {len(df)} rows[/green]")
        console.print(f"[blue]Columns: {', '.join(df.columns)}[/blue]\n")

        # Group by room_sku and division
        swaps = {}

        for idx, row in df.iterrows():
            room_sku = row.get('room_sku')
            division = self.normalize_division(row.get('division') or row.get('site'))
            swap_out_str = row.get('concatenated_itemqty')
            swap_in_sku = row.get('item_sku')

            if not room_sku or not division:
                console.print(f"[yellow]Row {idx+2}: Skipping - missing room_sku or division[/yellow]")
                continue

            # Parse swap out items
            swap_out_items = self.parse_concatenated_items(swap_out_str)

            # Create swap in item (assuming quantity 1 for each swap in item)
            # You may need to adjust this based on your actual data format
            swap_in_items = [{'sku': swap_in_sku, 'quantity': 1}] if swap_in_sku else []

            # Create key for grouping
            key = (room_sku, tuple(sorted([division])))

            if key not in swaps:
                swaps[key] = {
                    'room_sku': room_sku,
                    'divisions': list(key[1]),
                    'swap_out_items': [],
                    'swap_in_items': [],
                }

            # Add items to swap lists (avoiding duplicates)
            for item in swap_out_items:
                if item not in swaps[key]['swap_out_items']:
                    swaps[key]['swap_out_items'].append(item)

            for item in swap_in_items:
                if item not in swaps[key]['swap_in_items']:
                    swaps[key]['swap_in_items'].append(item)

        return list(swaps.values())

    def export_to_json(self, swaps: List[Dict], output_dir: Path) -> List[Path]:
        """
        Export swap operations to individual JSON files for traceability

        Returns:
            List of paths to the exported JSON files
        """
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_files = []

        for idx, swap in enumerate(swaps, 1):
            # Create individual JSON file for each room swap
            json_file = output_dir / f"room_item_swap_{timestamp}_{swap['room_sku']}_{idx:03d}.json"

            with open(json_file, 'w') as f:
                json.dump(swap, f, indent=2)

            json_files.append(json_file)

        console.print(f"[green]Exported {len(swaps)} swap operations to {len(json_files)} individual JSON files in: {output_dir}[/green]")
        return json_files

    def import_from_spreadsheet(
        self,
        file_path: Path,
        sheet_name: Optional[str] = None,
        dry_run: bool = True
    ) -> Dict:
        """
        Import room item swaps from spreadsheet

        Args:
            file_path: Path to spreadsheet file
            sheet_name: Excel sheet name (if applicable)
            dry_run: If True, don't actually execute swaps

        Returns:
            Dict with statistics
        """
        stats = {
            'processed': 0,
            'created': 0,
            'failed': 0,
            'validation_errors': 0,
            'work_request_ids': []
        }

        # Process spreadsheet
        swaps = self.process_spreadsheet(file_path, sheet_name)

        # Export to JSON - each swap gets its own file for traceability
        json_dir = file_path.parent / 'json'
        json_files = self.export_to_json(swaps, json_dir)

        # Validate all swaps
        console.print("\n[bold]Validating swap operations...[/bold]")
        validation_table = Table()
        validation_table.add_column("Room SKU", style="cyan")
        validation_table.add_column("Divisions", style="white")
        validation_table.add_column("Swap Out", style="yellow")
        validation_table.add_column("Swap In", style="green")
        validation_table.add_column("Status", style="white")

        valid_swaps = []

        for swap in swaps:
            is_valid, errors = self.validate_swap_data(swap)
            stats['processed'] += 1

            swap_out_str = ', '.join([f"{item['sku']}:{item['quantity']}" for item in swap['swap_out_items']])
            swap_in_str = ', '.join([f"{item['sku']}:{item['quantity']}" for item in swap['swap_in_items']])

            if is_valid:
                valid_swaps.append(swap)
                validation_table.add_row(
                    swap['room_sku'],
                    ', '.join(swap['divisions']),
                    swap_out_str,
                    swap_in_str,
                    "[green]✓ Valid[/green]"
                )
            else:
                stats['validation_errors'] += 1
                error_str = '; '.join(errors)
                validation_table.add_row(
                    swap['room_sku'],
                    ', '.join(swap.get('divisions', [])),
                    swap_out_str,
                    swap_in_str,
                    f"[red]✗ {error_str}[/red]"
                )

        console.print(validation_table)
        console.print(f"\n[bold]Validation Summary:[/bold]")
        console.print(f"  Total: {stats['processed']}")
        console.print(f"  Valid: {len(valid_swaps)}")
        console.print(f"  Invalid: {stats['validation_errors']}")

        if dry_run:
            console.print("\n[yellow]DRY RUN - Not executing swaps[/yellow]")
            stats['created'] = len(valid_swaps)
            return stats

        # Execute swaps
        console.print("\n[bold]Executing room item swaps...[/bold]")

        with Progress() as progress:
            task = progress.add_task("[cyan]Processing swaps...", total=len(valid_swaps))

            for swap in valid_swaps:
                try:
                    # Convert to API request format
                    swap_request = SwapRoomItemsRequest(
                        RoomSku=swap['room_sku'],
                        SwapOutRoomItems=[
                            RoomItem(Sku=item['sku'], Quantity=item['quantity'])
                            for item in swap['swap_out_items']
                        ],
                        SwapInRoomItems=[
                            RoomItem(Sku=item['sku'], Quantity=item['quantity'])
                            for item in swap['swap_in_items']
                        ],
                        Divisions=swap['divisions']
                    )

                    result = self.client.swap_room_items(swap_request)

                    if result:
                        stats['created'] += 1
                        if 'workrequest_id' in result:
                            wr_id = result['workrequest_id']
                            stats['work_request_ids'].append(wr_id)
                            console.print(f"[green]✓[/green] Room {swap['room_sku']}: Swap successful - Work Request ID: [cyan]{wr_id}[/cyan]")
                        else:
                            console.print(f"[green]✓[/green] Room {swap['room_sku']}: Swap successful (completed immediately)")
                    else:
                        stats['failed'] += 1
                        console.print(f"[red]✗[/red] Room {swap['room_sku']}: Swap failed")

                except Exception as e:
                    stats['failed'] += 1
                    console.print(f"[red]✗[/red] Room {swap['room_sku']}: Error - {e}")

                progress.advance(task)

        return stats


def main():
    """Main entry point for testing"""
    # This is for standalone testing
    from pathlib import Path

    api_url = os.getenv("API_URL", "http://127.0.0.1:8000")
    oauth_token_url = os.getenv("OAUTH_TOKEN_URL", "http://127.0.0.1:8010/token")
    oauth_client_id = os.getenv("OAUTH_CLIENT_ID", "m2m-api-client")
    oauth_client_secret = os.getenv("OAUTH_CLIENT_SECRET")

    if not oauth_client_secret:
        console.print("[red]OAUTH_CLIENT_SECRET not set[/red]")
        return

    oauth_config = OAuthConfig(
        client_id=oauth_client_id,
        client_secret=oauth_client_secret,
        token_url=oauth_token_url,
        use_m2m=True,
        use_pkce=False,
    )

    client = ECatalogAPIClient(api_url, oauth_config=oauth_config)
    client.authenticate_m2m()

    importer = RoomItemSwapImporter(client)

    # Example usage
    file_path = Path("data/room_item_swap/10-7-2025-FL-Active_Online_Rooms_with_CT_ET_Items_Inside.xlsx")

    if file_path.exists():
        stats = importer.import_from_spreadsheet(file_path, dry_run=True)
        console.print(f"\n[bold]Results:[/bold]")
        console.print(f"Processed: {stats['processed']}")
        console.print(f"Valid: {stats['created']}")
        console.print(f"Failed: {stats['failed']}")
        console.print(f"Validation Errors: {stats['validation_errors']}")


if __name__ == "__main__":
    main()
